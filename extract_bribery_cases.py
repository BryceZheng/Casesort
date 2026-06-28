#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
商业贿赂案例提取器 v2.1 — I列由 Claude Haiku 生成结构化摘要
数据来源：威科先行（PDF）+ 威科先行补充数据（PDF）+ 企查查（Excel）
I列（主要事实）使用 Claude Haiku API 生成结构化摘要
"""

import os, re, glob, sys, subprocess, shutil
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==================== AI 摘要（通过 claude CLI） ====================
# 使用已登录的 claude CLI 生成摘要，无需额外 API Key
_CLAUDE_CLI = shutil.which('claude')
USE_AI_SUMMARY = bool(_CLAUDE_CLI)

# ==================== 路径配置 ====================
PDF_BASE_DIR  = os.path.expanduser('~/Desktop/商业贿赂案例数据/威科先行数据')
WK_SUPP_DIR   = os.path.expanduser('~/Desktop/商业贿赂案例数据/威科先行补充数据')
QCC_DIR       = os.path.expanduser('~/Desktop/商业贿赂案例数据/企查查数据')
OUTPUT_2026       = os.path.expanduser('~/Desktop/商业贿赂案例库2026年版.xlsx')
OUTPUT_PRIOR      = os.path.expanduser('~/Desktop/商业贿赂案例库(2024-2025年).xlsx')
ORGANIZED_DIR     = os.path.expanduser('~/Desktop/最终挑选商业贿赂案例')

# ==================== 商业贿赂筛选 ====================

BRIBERY_CONFIRM = [
    r'商业贿赂',
    r'贿赂交易相对方',
    r'贿赂利用职权',
    r'第七条.{0,80}(贿赂|行贿)',
    r'(贿赂|行贿).{0,80}第七条',
    r'第八条.{0,80}(贿赂|行贿)',
    r'(贿赂|行贿).{0,80}第八条',
    r'第十九条.{0,80}(贿赂|行贿)',
    r'第二十四条.{0,80}(贿赂|行贿)',
    r'020203',
    r'以财物.{0,5}贿赂',
    r'给予.{0,25}好处费.{0,20}谋取',
    r'给予.{0,25}回扣.{0,20}谋取',
    r'给予.{0,25}带客.{0,10}费',
    r'带客回扣',
    r'谋取交易机会.{0,40}(贿赂|行贿|给予)',
    r'(行贿|贿赂).{0,40}谋取交易',
    r'向.{2,30}行贿',
    r'免费.{0,20}(设备|仪器|分析仪).{0,30}耗材',
    r'介绍费.{0,30}(谋取|拉拢|介绍)',
]

# 疑似商业贿赂：引用反法商业贿赂处罚条款（第19/24条），但正文未明确载明"贿赂"
# 第十九条（2019）和第二十四条（2025）是商业贿赂的专属处罚条款，引用即高度相关
# 注意：不能仅凭第七/八条推断（2019版第八条=虚假宣传，2025版第七条=虚假宣传）
BRIBERY_SUSPECTED = [
    # 引用了商业贿赂专属处罚条款（第十九/二十四条）
    r'依据[\s\S]{0,80}反不正当竞争法[\s\S]{0,100}第十九条',
    r'依据[\s\S]{0,80}反不正当竞争法[\s\S]{0,100}第二十四条',
    r'反不正当竞争法[\s\S]{0,80}第十九条[\s\S]{0,80}(没收违法所得|罚款)',
    r'反不正当竞争法[\s\S]{0,80}第二十四条[\s\S]{0,80}(没收违法所得|罚款)',
    # 违反了商业贿赂禁止条款本身（第七条=2019版, 第八条=2025版）
    # 虚假宣传组合（2019: 第八条+第二十条 / 2025: 第七条+第二十三条）已在 BRIBERY_EXCLUDE 中排除
    r'违反[\s\S]{0,80}反不正当竞争法[\s\S]{0,80}第[七八]条',
]

BRIBERY_EXCLUDE = [
    r'以欺骗.{0,3}贿赂等不正当手段取得.{0,10}(登记|许可)',
    r'骗取.{0,5}(商事|行政)登记',
    r'不起诉.{0,30}商事登记',
    r'刷单.{0,20}贿赂',
    r'出借.{0,5}营业执照',
    r'收取贿赂款合计',
    # 2025版第七条≠商业贿赂（是虚假宣传），处罚条款是第二十三条
    r'第七条.{0,400}第二十三条',
    r'第二十三条.{0,400}第七条',
    # 2019版第八条=虚假宣传，第二十条是其处罚条款；两者同文出现≠商业贿赂
    r'第八条.{0,300}第二十条',
    r'第二十条.{0,300}第八条',
    # 执法结论语言：认定为虚假宣传（而非法条引用）
    r'(构成|涉嫌|属于|认定为).{0,30}虚假.{0,5}(的)?商业宣传',
    r'(构成|涉嫌|认定为).{0,20}虚假宣传',
    # 混淆行为（第六条）：引人误以为/误认为是他人商品/存在特定联系
    r'引人误(以为|认为).{0,50}(是他人|存在特定联系|有特定关联)',
    r'(混淆行为|实施混淆|商业混淆)',
    r'(擅自使用|使用与).{0,30}(相同或近似|知名|他人).{0,30}(名称|标志|包装|装潢|商品)',
    # 商标侵权案（由商标法处理，非反不正当竞争法商业贿赂）
    r'侵犯.{0,10}注册商标专用权',
    # 虚假宣传案案件名称中明确标注（案例标题/文书标题含"虚假宣传"）
    r'(虚假宣传|引人误解.{0,5}(虚假宣传|商业宣传))案',
    # 药品经营/生产质量合规违规（GSP/GMP 违规），与商业贿赂无关
    r'未.{0,3}(遵守|执行|落实).{0,5}(GSP|GMP|药品经营质量管理规范|药品生产质量管理规范)',
]

_SUSPECTED_REMARK = '疑似商业贿赂（适用反法第十九条/第二十四条，文书未明确载明贿赂行为，请人工核查）'


def is_bribery_case(text):
    for pat in BRIBERY_EXCLUDE:
        if re.search(pat, text):
            return False
    for pat in BRIBERY_CONFIRM:
        if re.search(pat, text):
            return True
    return False


def is_suspected_bribery(text):
    """引用反法第十九/二十四条（商业贿赂专属处罚条款），但正文未明确写出"贿赂" → True"""
    for pat in BRIBERY_EXCLUDE:
        if re.search(pat, text):
            return False
    if is_bribery_case(text):
        return False
    return any(re.search(pat, text) for pat in BRIBERY_SUSPECTED)


# ==================== 字段提取 ====================

def _get_header_value(text, field):
    patterns = {
        '处罚机构': [
            r'处\s*罚\s*机\s*构[：:]\s*(.+?)[\n\r]',
            r'执\s*法\s*部\s*门[：:]\s*(.+?)[\n\r]',
            r'作出行政处罚决定机关名称\s*(.+?)[\n\r]',
        ],
        '案号': [
            r'发\s*文\s*案\s*号[：:]\s*(.+?)[\n\r]',
            r'行政处罚决定书[文号]*[：:]\s*(.{5,50}?)[\n\r]',
        ],
        '处罚时间': [
            r'处\s*罚\s*时\s*间[：:]\s*(\d{4}[.\-年]\d{1,2}[.\-月]?\d{0,2})',
            r'作出.{0,10}处罚.{0,5}日期[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{0,2})',
        ],
    }
    for pat in patterns.get(field, []):
        m = re.search(pat, text)
        if m:
            return m.group(1).strip()
    return None


def extract_case_number(text):
    val = _get_header_value(text, '案号')
    if val:
        return val
    m = re.search(r'[（(〔\[［][^\n〕\]）)]{0,50}[\d]{4}[^\n〕\]）)]{0,30}[〕\]）)][^\n]{0,20}号', text[:600])
    if m:
        return m.group(0).strip()
    return '/'


def normalize_date(raw):
    if not raw:
        return '/'
    raw = raw.strip()
    m = re.match(r'(\d{4})[.\-/年](\d{1,2})(?:[.\-/月](\d{0,2}))?', raw)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), (m.group(3) or '').zfill(2)
        if d and d != '00':
            return f'{y}.{mo}.{d}'
        return f'{y}.{mo}'
    return raw


def extract_punishment_time(text):
    raw = _get_header_value(text, '处罚时间')
    if raw:
        return normalize_date(raw)
    m = re.search(r'(\d{4})[-./年](\d{1,2})[-./月](\d{1,2})', text[:2000])
    if m:
        return f'{m.group(1)}.{m.group(2).zfill(2)}.{m.group(3).zfill(2)}'
    return '/'


def extract_institution(text):
    val = _get_header_value(text, '处罚机构')
    if val:
        return val.strip()
    m = re.search(r'处\s*罚\s*机\s*关\s*[\n\r]\s*(.{4,30}市场监督)', text)
    if m:
        return m.group(1).strip()
    return '/'


_CASE_TYPE_SUFFIXES = [
    '贿赂交易相对方的工作人员案',
    '贿赂利用职权或者影响力影响交易的单位或者个人案',
    '商业贿赂案',
    '涉嫌商业贿赂案',
    '关于',
]

_CASE_NUMBER_RE = re.compile(r'[〔〕\[\]（）]|\d{4}[^\d]{0,5}号')


def _clean_case_name_from_filename(filename_no_ext):
    name = filename_no_ext
    if _CASE_NUMBER_RE.search(name):
        return None
    for suffix in _CASE_TYPE_SUFFIXES:
        name = name.replace(suffix, '')
    name = re.sub(r'^[（(（\s]*(?:局审)[）)）\s]*', '', name).strip()
    name = name.strip('（）()【】 \t')
    m = re.search(r'^(.{3,30}?)(?:贿赂|涉嫌|商业贿赂)', name)
    if m:
        name = m.group(1).strip()
    return name if len(name) >= 2 else None


def extract_company_name(text, filename):
    # 方法1: "当事人：XXX" 格式（最可靠）
    m = re.search(r'当事人[：:]\s*([^\n（(【]{3,50})(?:\n|（|主体|统一社会|法定)', text)
    if m:
        name = re.sub(r'\s+', '', m.group(1)).strip('。，、 ')
        if len(name) >= 3:
            return name

    # 方法2: 表格格式"被处罚单位/人"
    m = re.search(
        r'被\s*\n\s*处\s*\n\s*罚\s*\n\s*(?:单\s*\n\s*位\s*\n\s*)?(?:[（(]\s*\n\s*被\s*\n\s*处\s*\n\s*罚\s*\n\s*人\s*\n\s*[）)]\s*\n\s*)?([^\n\d（(【][^\n]{2,50})\s*\n',
        text
    )
    if m:
        name = m.group(1).strip()
        if len(name) >= 2 and '市场监督' not in name:
            return name

    # 方法3: 文件名（清理后）
    base_name = os.path.splitext(filename)[0]
    cleaned = _clean_case_name_from_filename(base_name)
    if cleaned:
        return cleaned

    # 方法4: 文本第一行（若不是标题性内容）
    first_line = text.split('\n')[0].strip()
    if (len(first_line) >= 2
            and '处罚' not in first_line
            and '贿赂' not in first_line
            and '案' not in first_line
            and '〔' not in first_line
            and not re.search(r'\d{4}', first_line)):
        return first_line

    # 文件名含案号格式时不返回文件名
    if _CASE_NUMBER_RE.search(base_name):
        return '/'
    return base_name or '/'


# ==================== 被行贿方提取 ====================

# 被行贿方必须包含的实体类型关键词（至少其中之一）
_BRIBEE_ENTITY_WORDS = (
    '公司', '医院', '诊所', '卫生院', '卫生室', '医疗', '药房', '中心',
    '工作人员', '采购', '院长', '医生', '村医', '护士', '主任', '科室',
    '负责人', '经理', '司机', '员工', '职工', '干部', '官员',
)

# 无效的被行贿方文本（片语/法条描述/支付对象）
_BRIBEE_INVALID = (
    '下列', '上述', '有关单位', '单位或者个人', '等人', '等单位',
    '烟酒', '礼品', '财物', '现金', '好处费', '回扣', '介绍费', '钱款',
    '罚款', '没收', '警告', '改正',
    '交易相对方', '影响交易', '利用职权', '相关事务',
    '市场监督', '违法行为', '供应商', '竞争优势',
)


def _valid_bribee(s):
    s = s.strip().rstrip('，。的：: ')
    if len(s) < 2 or len(s) > 40:
        return False
    if any(kw in s for kw in _BRIBEE_INVALID):
        return False
    # 必须含有实体类型词，或看起来像人名（2-4汉字）
    has_entity = any(kw in s for kw in _BRIBEE_ENTITY_WORDS)
    looks_like_person = bool(re.match(r'^[一-鿿]{2,4}$', s))
    return has_entity or looks_like_person


def extract_bribee(text):
    flat = re.sub(r'\s+', ' ', text)

    patterns = [
        # 向 [entity] 行贿/支付
        r'向\s*(.{2,35}?(?:公司|医院|诊所|卫生院|医疗机构|中心|工作人员|医生|村医|院长|护士|主任|经理|负责人|采购|药房))\s*(?:行贿|贿赂|支付好处费|给予财物|给予钱款|支付回扣)',
        # 给予 [entity] [物品/钱款]
        r'给予\s*(.{2,35}?(?:公司|医院|诊所|卫生院|医疗机构|工作人员|医生|村医|院长|护士|主任|经理|负责人|采购|司机))\s*(?:钱款|财物|现金|好处费|回扣|礼品|佣金|介绍费)',
        # 拉拢 [entity]
        r'拉拢\s*(.{2,30}?(?:工作人员|负责人|院长|主任|医生|经理|采购))',
        # 给 [entity] 带客回扣/好处费
        r'给\s*(.{3,30}?(?:医生|村医|医院|诊所|负责人|司机|工作人员|采购))\s*(?:带客回扣|回扣|好处费|介绍费)',
        # 医疗器械向医院模式
        r'免费.{0,10}为\s*(.{3,40}?(?:医院|机构|诊所|卫生院))\s*(?:增加|配备|提供)',
        r'为\s*(.{3,40}?(?:医院|机构|诊所))\s*(?:唯一|独家|签订)',
        r'(?:销售|配送|供货|供应).{0,20}给\s*(.{3,40}?(?:医院|机构|诊所))',
        # 贿赂/行贿 [entity]
        r'(?:贿赂|行贿)\s*(.{2,30}?(?:公司|医院|诊所|机构|工作人员|负责人|院长|主任))\s*(?:共计|合计|元[，。]|以谋)',
    ]

    for pat in patterns:
        m = re.search(pat, flat)
        if m:
            bribee = m.group(1).strip().rstrip('，。的')
            if _valid_bribee(bribee):
                return bribee

    return '未披露'


# ==================== 被行贿方角色 ====================

def classify_bribee_role(text):
    flat = re.sub(r'\s+', ' ', text)

    hospital_patterns = [
        r'向.{0,40}(医院|诊所|卫生院|医疗机构|整骨医院|村医|科室|检验科|处方).{0,40}(行贿|贿赂|给予|支付|免费)',
        r'免费.{0,10}为.{0,40}(医院|诊所|卫生院|医疗机构).{0,20}(增加|配备|提供)',
        r'给予.{0,10}(医院|诊所|村医|科室|处方权).{0,20}(回扣|好处费|介绍费|财物)',
        r'(医院|诊所|卫生院).{0,10}(工作人员|医生|护士|科室|负责人).{0,20}(回扣|好处费|介绍费)',
    ]
    for pat in hospital_patterns:
        if re.search(pat, flat):
            return '医院'

    has_art3 = bool(
        re.search(r'第[七八]条.{0,80}第.{0,3}[（(]三[）)].{0,3}项', flat) or
        re.search(r'利用.{0,15}(职权|影响力).{0,15}影响交易', flat)
    )
    has_art1 = bool(re.search(r'第[七八]条.{0,80}第.{0,3}[（(]一[）)].{0,3}项', flat))
    has_art2 = bool(re.search(r'第[七八]条.{0,80}第.{0,3}[（(]二[）)].{0,3}项', flat))

    if has_art3:
        return '利用职权或者影响力影响交易的单位或者个人'
    if has_art1 or has_art2:
        return '交易相对方的工作人员'

    if re.search(r'利用.{0,15}(职权|影响力)', flat):
        return '利用职权或者影响力影响交易的单位或者个人'
    if re.search(r'(工作人员|采购人员|员工|负责人|主任|科长).{0,30}(行贿|贿赂|回扣|好处费)', flat):
        return '交易相对方的工作人员'
    if re.search(r'向交易相对方', flat):
        return '交易相对方的工作人员'
    if re.search(r'贿赂.{0,10}交易相对方.{0,10}工作人员', flat):
        return '交易相对方的工作人员'

    return '未披露'


# ==================== 行贿手段 ====================

def classify_bribe_method(text):
    flat = re.sub(r'\s+', ' ', text)
    methods = []

    if re.search(r'(?:捆绑|搭售|只能在当事人处购买|配套耗材只能|耗材.*只能.*购买)', flat):
        methods.append('捆绑销售耗材')

    if re.search(r'(?:低价|优惠价|免费|无偿).{0,30}(?:销售|提供|赠送|增加|配备).{0,30}(?:设备|仪器|分析仪|检测仪|机器)', flat):
        methods.append('低价销售设备')

    if re.search(r'优先.{0,15}(?:销售|供应|配送|提供).{0,15}耗材', flat):
        methods.append('优先销售耗材')

    if re.search(r'(?:回扣|带客回扣|返点|佣金|介绍费|提成)', flat):
        methods.append('返点')

    if re.search(r'(?:现金|钱款|转账|汇款|微信.*转账|支付宝|线上转账)', flat):
        if not methods or '返点' not in methods:
            methods.append('现金')

    if re.search(r'(?:以财物|礼品|礼卡|购物卡|有价证券|实物|礼物)', flat):
        if not methods:
            methods.append('财物')

    if not methods and re.search(r'以财物.{0,5}贿赂', flat):
        methods.append('财物')

    return '、'.join(methods) if methods else '未披露'


# ==================== 行业分类 ====================

INDUSTRY_RULES_BY_NAME = [
    ('医疗器械',    ['医疗器械', '器械有限', '医疗设备', '医疗仪器', '医疗科技', '医疗技术有限', '康复器具']),
    ('医疗机构',    ['医院', '诊所', '卫生院', '医疗中心', '医疗集团', '卫生室', '卫生所', '门诊']),
    ('医药/制药',   ['药业', '药品', '药房', '医药', '制药', '大药房', '药店', '生物医药', '医药集团']),
    ('新能源',      ['新能源', '光伏', '风电', '储能', '清洁能源', '绿色能源']),
    ('石油/能源',   ['石化', '石油', '燃气', '天然气', '煤炭', '煤矿', '矿业']),
    ('汽车',        ['汽车', '汽修', '机动车', '检测站', '年检', '车辆检测', '整车', '汽配', '4s', '二手车', '新车']),
    ('旅游',        ['旅游', '旅行社', '景区', '旅行', '旅游开发']),
    ('餐饮',        ['餐厅', '餐饮', '食堂', '饭店', '酒楼', '外卖', '美食', '饮食', '酒家', '小吃']),
    ('住宿/酒店',   ['酒店', '宾馆', '民宿', '客栈']),
    ('食品',        ['食品', '饮料', '农产品', '粮油', '副食', '食品加工', '饮品']),
    ('建筑/工程',   ['建设', '建筑', '施工', '装饰', '装修', '路桥', '基建', '市政', '工程局', '建工', '装潢']),
    ('教育/培训',   ['教育', '学校', '培训', '幼儿园', '技能', '职业学校', '培训中心', '教学']),
    ('物流/运输',   ['物流', '快递', '运输', '货运', '航运', '配送', '仓储', '快运']),
    ('包装材料',    ['包装', '纸箱', '包装材料', '包材', '彩印包装']),
    ('印刷',        ['印刷', '印务', '印刷厂', '印业']),
    ('家具/木业',   ['家具', '木业', '木材', '板材', '木制品', '定制家居']),
    ('金属/钢材',   ['金属', '钢材', '铝业', '铝合金', '钢管', '五金', '不锈钢', '钢铁', '铜业', '铝制']),
    ('化工',        ['化工', '试剂', '化学品', '化学试剂', '耗材有限', '实验室']),
    ('机械/设备',   ['机械', '数控', '机床', '工业设备', '自动化', '机器设备', '机器人']),
    ('电子/电气',   ['电气', '电子', '电力', '电缆', '电器', '电机', '电控', '芯片', '半导体']),
    ('纺织/服装',   ['纺织', '服装', '面料', '织物', '服饰', '布料']),
    ('建材/装饰',   ['石材', '砖石', '建材', '装饰材料', '涂料', '瓷砖', '地板', '保温材料', '门窗']),
    ('科技/软件',   ['科技', '软件', '信息技术', '互联网', '云计算', '网络技术', '数字科技', '人工智能',
                     'AI有限', '数据有限', '信息有限', '网络有限']),
    ('金融/保险',   ['银行', '保险', '金融', '基金', '证券', '担保', '小额贷款']),
    ('房地产/中介', ['房地产', '地产', '置业', '开发有限公司', '房产经纪', '中介', '评估']),
    ('商贸/批发',   ['商贸', '贸易', '商行', '超市', '零售', '进出口', '批发', '商业']),
    ('医疗服务',    ['非急救', '转运', '康复', '护理院', '养老院', '养老服务']),
    ('广告/媒体',   ['广告', '传媒', '媒体', '文化传播', '影视', '文化有限', '文创']),
    ('采矿/能源',   ['矿业', '矿产', '采矿', '煤矿', '铁矿', '采掘']),
]


def classify_industry(company_name, text):
    name_lower = (company_name or '').lower()

    for industry, keywords in INDUSTRY_RULES_BY_NAME:
        for kw in keywords:
            if kw in company_name or kw.lower() in name_lower:
                return industry

    medical_device_override = bool(re.search(
        r'(医用耗材|检验耗材|骨科|心内科|医疗耗材|手术器械|介入耗材|影像|内镜|超声|放射)',
        text[:1500]
    ))
    if medical_device_override:
        return '医疗器械'

    text_head = text[:800]
    text_head_clean = re.sub(
        r'(?:给予|向|贿赂|行贿|支付).{0,5}(医院|诊所|卫生院|医生|村医|院长|司机|旅行社).{0,100}', '', text_head
    )
    for industry, keywords in INDUSTRY_RULES_BY_NAME:
        for kw in keywords:
            if kw in text_head_clean:
                return industry

    return '其他'


# ==================== 金额提取 ====================

CN_DIGIT = {'零': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
            '六': 6, '七': 7, '八': 8, '九': 9,
            '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5,
            '陆': 6, '柒': 7, '捌': 8, '玖': 9,
            '十': 10, '拾': 10, '百': 100, '佰': 100,
            '千': 1000, '仟': 1000, '万': 10000, '亿': 100000000}


def _parse_cn_amount(s):
    s = s.strip()
    m = re.match(r'^(\d+(?:\.\d+)?)万?$', s)
    if m:
        val = float(m.group(1))
        if '万' in s:
            val *= 10000
        return int(val)
    s2 = s.replace(',', '').replace('，', '')
    try:
        return int(float(s2))
    except ValueError:
        pass
    result = 0
    tmp = 0
    unit = 1
    for ch in reversed(s):
        if ch in CN_DIGIT:
            v = CN_DIGIT[ch]
            if v >= 10:
                if v > unit:
                    unit = v
                    if v == 10000:
                        result += tmp
                        tmp = 0
                else:
                    tmp *= v
            else:
                tmp += v * unit
                unit = 1
    result += tmp
    return result if result > 0 else None


def _extract_amount_from_text(text, patterns):
    text_flat = re.sub(r'\s+', ' ', text)
    for pat in patterns:
        m = re.search(pat, text_flat)
        if m:
            raw = m.group(1).strip()
            raw = raw.replace('¥', '').replace('￥', '').strip()
            # Bug fix: use text_flat (not text) for post-match window
            post = text_flat[m.end():m.end() + 4]
            is_wan = '万' in raw or '万元' in post
            raw = re.sub(r'[万元]', '', raw).strip()
            val = _parse_cn_amount(raw)
            if val and val > 0:
                return int(val * 10000) if is_wan and val < 10000 else val
    return None


def extract_bribe_amount(text):
    patterns = [
        r'(?:行贿|贿赂|给予.{0,15}共计|支付.{0,15}共计).{0,30}?(?:人民币|共计)?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
        r'(?:好处费|回扣|介绍费|佣金).{0,20}?(?:共计|合计|共)?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
        r'给予.{2,30}?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
    ]
    val = _extract_amount_from_text(text, patterns)
    return val if val else '/'


def extract_fine_amount(text):
    flat = re.sub(r'\s+', ' ', text)
    patterns = [
        r'罚款(?:人民币)?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
        r'处(?:罚款)?(?:人民币)?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
        r'罚款金额.{0,5}(\d+(?:\.\d+)?)\s*(?:万元|（万元）)',
    ]
    val = _extract_amount_from_text(flat, patterns)
    if not val:
        m = re.search(r'罚款(\d+(?:\.\d+)?)元', flat)
        if m:
            return int(float(m.group(1)))
        m = re.search(r'罚款金额.{0,3}(\d+\.\d+)', flat)
        if m:
            return int(float(m.group(1)) * 10000)
        m = re.search(r'(?:罚款|处罚).{0,60}(?:￥|¥)\s*(\d[\d,.]+)', flat)
        if m:
            return int(float(m.group(1).replace(',', '')))
        m = re.search(r'\d\.\s*罚款(\d+(?:\.\d+)?)元', flat)
        if m:
            return int(float(m.group(1)))
    return val if val else '/'


def extract_confiscated_amount(text):
    patterns = [
        r'没收违法所得(?:人民币)?([零一二三四五六七八九十百千万亿壹贰叁肆伍陆柒捌玖拾佰仟\d,，.]+)(?:元|万元)',
        r'没收违法所得.{0,10}(\d+(?:\.\d+)?)(?:元|万元)',
    ]
    val = _extract_amount_from_text(text, patterns)
    if not val:
        m = re.search(r'没收.{0,5}违法所得.{0,20}(\d+\.\d+)\s*(?:、|万元|\（万元）|\n)', text)
        if m:
            return int(float(m.group(1)) * 10000)
    return val if val else '/'


# ==================== 地域提取 ====================

PROVINCES = [
    '北京市', '天津市', '上海市', '重庆市',
    '河北省', '山西省', '辽宁省', '吉林省', '黑龙江省',
    '江苏省', '浙江省', '安徽省', '福建省', '江西省',
    '山东省', '河南省', '湖北省', '湖南省', '广东省',
    '海南省', '四川省', '贵州省', '云南省', '陕西省',
    '甘肃省', '青海省',
    '内蒙古自治区', '广西壮族自治区', '西藏自治区',
    '宁夏回族自治区', '新疆维吾尔自治区',
]

CITY_TO_PROVINCE = {
    '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
    '深圳': '广东省', '广州': '广东省', '东莞': '广东省', '佛山': '广东省',
    '珠海': '广东省', '汕头': '广东省', '惠州': '广东省', '中山': '广东省',
    '肇庆': '广东省', '清远': '广东省',
    '杭州': '浙江省', '宁波': '浙江省', '温州': '浙江省', '嘉兴': '浙江省',
    '慈溪': '浙江省', '淳安': '浙江省', '绍兴': '浙江省', '台州': '浙江省',
    '义乌': '浙江省', '金华': '浙江省', '舟山': '浙江省', '丽水': '浙江省',
    '东阳': '浙江省', '长兴': '浙江省', '瑞安': '浙江省', '海宁': '浙江省',
    '苏州': '江苏省', '南京': '江苏省', '无锡': '江苏省', '常州': '江苏省',
    '常熟': '江苏省', '南通': '江苏省', '盐城': '江苏省', '徐州': '江苏省',
    '扬州': '江苏省', '镇江': '江苏省', '泰州': '江苏省', '淮安': '江苏省',
    '溧阳': '江苏省', '宜兴': '江苏省', '海门': '江苏省',
    '成都': '四川省', '武汉': '湖北省', '西安': '陕西省', '长沙': '湖南省',
    '郑州': '河南省', '沈阳': '辽宁省', '大连': '辽宁省',
    '合肥': '安徽省', '寿县': '安徽省', '六安': '安徽省', '石台': '安徽省',
    '池州': '安徽省', '黄山': '安徽省', '芜湖': '安徽省', '蚌埠': '安徽省',
    '南昌': '江西省', '九江': '江西省', '抚州': '江西省',
    '赣州': '江西省', '吉安': '江西省', '新余': '江西省',
    '福州': '福建省', '厦门': '福建省', '漳州': '福建省', '泉州': '福建省',
    '三明': '福建省', '龙岩': '福建省',
    '济南': '山东省', '青岛': '山东省', '烟台': '山东省', '济宁': '山东省',
    '临沂': '山东省', '潍坊': '山东省', '菏泽': '山东省', '淄博': '山东省',
    '太原': '山西省', '石家庄': '河北省', '保定': '河北省', '唐山': '河北省',
    '哈尔滨': '黑龙江省', '长春': '吉林省',
    '南宁': '广西壮族自治区', '桂林': '广西壮族自治区',
    '昆明': '云南省', '曲靖': '云南省', '玉溪': '云南省',
    '贵阳': '贵州省', '遵义': '贵州省', '毕节': '贵州省',
    '呼和浩特': '内蒙古自治区', '包头': '内蒙古自治区', '赤峰': '内蒙古自治区',
    '兰州': '甘肃省', '西宁': '青海省',
    '银川': '宁夏回族自治区',
    '乌鲁木齐': '新疆维吾尔自治区', '吐鲁番': '新疆维吾尔自治区', '喀什': '新疆维吾尔自治区',
    '拉萨': '西藏自治区', '海口': '海南省', '三亚': '海南省',
    '沅陵': '湖南省', '张家界': '湖南省', '常德': '湖南省',
    '株洲': '湖南省', '湘潭': '湖南省', '邵阳': '湖南省',
    '保康': '湖北省', '宜昌': '湖北省', '荆州': '湖北省', '鄂州': '湖北省', '十堰': '湖北省',
    '新郑': '河南省', '许昌': '河南省', '洛阳': '河南省', '南阳': '河南省',
    '广元': '四川省', '泸州': '四川省', '德阳': '四川省', '南充': '四川省',
    '宜宾': '四川省', '绵阳': '四川省', '眉山': '四川省', '乐山': '四川省', '遂宁': '四川省',
    '延安': '陕西省', '宝鸡': '陕西省', '咸阳': '陕西省',
}


def extract_region(institution, text):
    combined = (institution or '') + text[:1000]
    for prov in PROVINCES:
        if prov in combined[:200]:
            return prov
        short = prov[:2]
        if institution and short in institution:
            return prov

    for city, prov in CITY_TO_PROVINCE.items():
        if institution and city in institution:
            return prov
        if city in text[:500]:
            return prov

    m = re.search(r'住\s*所.{0,5}\n\s*(.{4,50})(?=\n)', text)
    if m:
        addr = m.group(1)
        for prov in PROVINCES:
            if prov in addr or prov[:2] in addr:
                return prov
        for city, prov in CITY_TO_PROVINCE.items():
            if city in addr:
                return prov

    return '/'


# ==================== 主要事实提取 ====================

# 需要跳过的程序性语句（不属于违法事实叙述）
SKIP_SENTENCE_PATTERNS = [
    r'^本局.{0,10}(决定|认定|查明|发出)',
    r'^依据.{0,5}(《|法|第)',
    r'^根据.{0,5}(《|法|第)',
    r'^上述事实',
    r'证据.{0,5}证明$',
    r'^认定为初次违法',
    r'^违法所得无法计算',
    r'^责令.{0,5}改正',
    r'罚款.{1,10}元[。，]?$',
    r'^当事人.{0,5}(自收到|逾期)',
    r'^综上',
    r'^经查',
    r'^现场检查(发现)?',
    r'^我局.{0,10}(执法|检查|收到)',
    r'^决定行政处罚',          # "决定行政处罚种类、依据、内容：……"
    r'^处罚(种类|依据|结果|内容)[：:]',
    r'^行政处罚(种类|依据|决定|履行)',
    r'行政处罚履行方式',       # "行政处罚履行方式和期限：当事人自收到……"
    r'自收到.{0,10}处罚决定书.{0,5}之日起',
    r'^扫一扫',               # PDF 二维码区域 OCR 残留
    r'手机阅读更方便',
    r'(执行标准|使用证书编号|产品类型|产品原料|产品名称|适宜人群|豫健用准|适用标准)',
    r'(询问笔录|现场检查笔录|营业执照复印件|当事人陈述|法定代表人身份证)',
]

# OCR 碎片行检测：满足任一规则的行直接丢弃
_GARBAGE_LINE_RE = re.compile(
    r'^[\dA-Za-z]{1,5}$'                          # 孤立编码如 "9T"、"36"
    r'|^[\d]{1,4}[/／]\d+$'                       # 页码如 "1/5"
    r'|执行标准|使用证书编号|产品类型|产品名称'
    r'|适宜人群|适用标准|产品原料'
    r'|统一社会信用代码\s*[A-Z0-9]{15,}'          # 信用代码行
    r'|身份证.{0,5}(号码|编号)[：:]'              # 身份证号码行（OCR常与年份粘连）
    r'|\d+[号层幢栋楼室]\s*[\d,，]'              # 地址碎片（含后续数字）
    r'|询问笔录|现场检查笔录|营业执照复印件'
    r'|当事人陈述|法定代表人身份证'
    r'|工商登记资料|现场照片|证据目录'
)

# 句子行贿叙事相关度打分关键词
_FACT_SCORE_PATTERNS = [
    (3, r'当事人.{0,40}(为了|为谋取|为获取).{0,40}(机会|利益|订单|资格|优势)'),
    (3, r'向.{2,35}(行贿|贿赂|支付好处费|给予财物|给予回扣)'),
    (3, r'给予.{2,35}(好处费|回扣|介绍费|财物|现金).{0,20}(共计|合计|元)'),
    (2, r'(共计|合计).{0,10}[\d万]'),
    (2, r'(回扣|好处费|介绍费).{0,15}[\d万元]'),
    (2, r'贿赂.{2,30}(工作人员|采购|院长|医生|负责人|司机)'),
    (2, r'(免费|低价).{0,20}(设备|仪器|分析仪).{0,20}耗材'),
    (1, r'谋取.{0,20}(交易机会|竞争优势|业务|合同)'),
    (1, r'(当事人|经营者).{0,20}(为了|在).{0,20}(销售|推广|承接)'),
    (1, r'(行贿|贿赂|好处费|回扣|介绍费|带客费)'),
]


def _score_sentence(s):
    score = 0
    for weight, pat in _FACT_SCORE_PATTERNS:
        if re.search(pat, s):
            score += weight
    return score


def _clean_label_fragments(text):
    """去除表格格式中穿插的单字标签行"""
    lines = text.split('\n')
    cleaned = []
    for line in lines:
        s = line.strip()
        if len(s) <= 2 and re.match(r'^[一-鿿]+$', s):
            continue
        cleaned.append(line)
    return '\n'.join(cleaned)


_AI_SUMMARY_PROMPT = """以下是一份商业贿赂行政处罚决定书的文本（含OCR识别，可能有少量噪音）。
请根据文中的具体违法事实，用1-3句话简洁总结。

【正确写法示例】（必须达到这个具体程度）：
- 当事人通过向鹰潭市**幼儿园提供幼教云相关软硬件设施供其无偿使用的方式贿赂幼儿园，以谋取交易机会或者竞争优势，收取幼教云服务费。
- 2018年5月至2023年2月，当事人为谋取交易机会，向某医院在职医生以"共同管理费"名义支付费用共5,120,503元，获得推介病例1148例。
- 当事人从事机动车检验代办业务，为让车辆优先通过年检，向长兴县某检测公司检验员给予好处费，以获取更多客户资源和交易机会。
- 当事人通过给予负责消防器材采购的村党委委员回扣，获取与太仓市双凤镇某村的交易机会，向其销售消防器材。
- 苏州明博机电有限公司为长期承接群胜科技（苏州）有限公司的维修项目，通过微信转账多次给予时任该公司五金加工科主任共计230,500元，谋取交易机会。
- 当事人自2021年起，以免费投放血凝分析仪等设备供医院无偿使用的方式贿赂泸州市纳溪区人民医院，以换取医院指定采购其耗材的交易机会。

【禁止写法】（以下写法无效，必须避免）：
✗ "当事人实施商业贿赂。"——没有说明对谁、用什么方式、谋取什么
✗ "当事人采用财物手段贿赂交易相对方工作人员。"——照抄法条，无具体细节
✗ "当事人存在商业贿赂违法行为，违法所得13824元。"——只有结论，没有事实
✗ "当事人违反了《反不正当竞争法》第七条……"——法律条文不是事实摘要

【格式规则】：
- 结构：行贿方（谁）→ 手段（通过/向…给予/提供）→ 被行贿对象（具体人/单位）→ 目的（谋取交易机会/竞争优势）
- 有金额必须写；有时间段必须写；有次数必须写；有被行贿人姓名/职务必须写
- 输出控制在250字以内，不加引号、不加任何说明前缀
- 以下情况直接输出"/"（无其他内容）：文本中只有法律条文/处罚决定/程序性陈述，完全没有具体行贿事实

行贿方：{company}
处罚文书文本：
{text}"""


_LAW_BOILERPLATE_PATTERNS = [
    r'^经营者采用财物手段贿赂交易相对方',
    r'^通过给予财物或其他手段.{0,5}贿赂交易相对方',
    r'^贿赂交易相对方的工作人员',
    r'^贿赂利用职权或者影响力影响交易的单位',
    # 引用法条定义文字（含 OCR 碎片如"第处八条"）
    r'违反.{0,80}反不正当竞争法.{0,80}第.{0,3}[七八]条.{0,50}经营者不得采用财物',
    r'经营者不得采用财物或者其他手段贿赂下列',
    r'违反本法第.{0,3}[七八]条规定贿赂他人.{0,30}由监督检查部门',
    r'有关单位违反本法第.{0,3}[七八]条规定贿赂他人或者收受贿赂',
    # 处罚决定书套语（无事实，仅处罚金额/依据）
    r'(我局|本局)决定.{0,20}(减轻|从轻|予以).{0,10}处罚',
    r'处罚如下.{0,5}[一１]',
]


def _is_law_boilerplate(text):
    if not text or text in ('/', '（原始处罚文书未载明违法事实详情）'):
        return False
    return any(re.search(p, text) for p in _LAW_BOILERPLATE_PATTERNS)


def _ai_summarize_fact(raw_text, company_name=''):
    """调用 claude CLI 生成结构化商业贿赂事实摘要。失败时返回 None。"""
    if not USE_AI_SUMMARY or not _CLAUDE_CLI:
        return None
    cleaned = re.sub(r'\s+', '', raw_text or '')
    if len(cleaned) < 40:
        return None
    prompt = _AI_SUMMARY_PROMPT.format(
        company=company_name or '当事人',
        text=raw_text[:3500],
    )
    try:
        result = subprocess.run(
            [_CLAUDE_CLI, '-p', prompt],
            capture_output=True, text=True, timeout=45,
        )
        if result.returncode == 0:
            output = result.stdout.strip().strip('"').strip()
            # 去掉 AI 可能添加的元注释段落（"（注：...）"）
            output = re.sub(r'\n*[（(]注[：:][\s\S]{0,300}[）)]', '', output).strip()
            # 若输出 "/" 或空，返回 None
            if not output or output == '/' or output.startswith('/\n') or output.startswith('/ '):
                return None
            if len(output) < 8:
                return None
            if _is_law_boilerplate(output):
                return None
            return output
        return None
    except Exception as e:
        print(f'  [AI] 摘要失败: {e}')
        return None


def _condense_fact(text, max_chars=250):
    """将事实文本压缩为可读摘要（去 OCR 碎片 + 按行贿相关度选句）"""
    if not text:
        return '/'

    # Step 1: 去单字表格标签
    text = _clean_label_fragments(text)

    # Step 2: 逐行过滤 OCR 碎片、产品规格、地址碎片
    clean_lines = []
    for line in text.split('\n'):
        stripped = line.strip()
        if stripped and not _GARBAGE_LINE_RE.search(stripped):
            clean_lines.append(stripped)
    text = ' '.join(clean_lines)

    # Step 3: 去除所有空白，保留有效字符
    text = re.sub(r'\s+', '', text)
    # 清除行内 OCR 粘连噪音：纯数字串（身份证/编号/页码等）粘在汉字中间
    text = re.sub(r'(?<=[一-鿿])\d{8,}(?=[一-鿿年])', '', text)  # 长数字串粘连
    text = re.sub(r'[^一-鿿\d，。！？；：、《》【】（）()%·—元万年月日a-zA-Z]', '', text)

    if not text:
        return '/'

    # 短文本直接返回
    if len(text) <= max_chars:
        return text

    # Step 4: 分句，并剥离句首 OCR 单字残留（如"摘"、"书"、"定"等表格边框字符）
    sentences = re.split(r'(?<=[。！？])', text)
    clean = []
    for s in sentences:
        s = s.strip()
        # 去掉句首 1-2 个孤立汉字（不构成词的 OCR 残留）
        s = re.sub(r'^[一-鿿]{1,2}(?=[一-鿿]{2})', '', s)
        if len(s) > 8:
            clean.append(s)
    sentences = clean

    if not sentences:
        return text[:max_chars] + '……'

    # Step 5: 过滤程序性语句，同时打分
    scored = []
    for s in sentences:
        if any(re.search(p, s) for p in SKIP_SENTENCE_PATTERNS):
            continue
        scored.append((s, _score_sentence(s)))

    if not scored:
        scored = [(s, _score_sentence(s)) for s in sentences]

    # Step 6: 保留得分最高的若干句（最多4句），按原顺序排列
    top_idx = set(
        i for i, _ in sorted(enumerate(scored), key=lambda x: -x[1][1])[:4]
    )
    result = [s for i, (s, _) in enumerate(scored) if i in top_idx]

    # 如果最高分都是0（无明确行贿词），退化为按顺序取前3句
    if all(sc == 0 for _, sc in scored):
        result = [s for s, _ in scored[:3]]

    # Step 7: 合并，超出截断
    joined = ''.join(result)
    if len(joined) > max_chars + 80:
        joined = joined[:max_chars] + '……'
    return joined or '/'


def _find_raw_fact_block(text):
    """从 PDF 文本中找出最相关的违法事实原始段落，返回 (raw_block, is_structured)。
    is_structured=True 表示找到了明确标注的事实字段；False 表示靠关键词猜测的块。"""

    # Pattern 1: "违法事实：" 字段（威科先行结构化格式，最可靠）
    m = re.search(
        r'(?:违\s*法\s*事\s*实|处\s*罚\s*事\s*实)[：:]\s*(.{30,}?)(?=处罚依据|处罚内容|处罚结果|$)',
        text, re.DOTALL
    )
    if m:
        return m.group(1), True

    # Pattern 2: "主要违法事实：" 摘要格式
    m = re.search(r'主要违法事实[：:]\s*(.{20,}?)(?=行政处罚种类|行政处罚履行|$)', text, re.DOTALL)
    if m:
        fact = m.group(1)
        fact = re.sub(r'(?m)^[行政处罚决定书（全文或摘要）]\s+', '', fact)
        return fact, True

    # Pattern 3: "调查认定的事实：" 全文格式
    m = re.search(r'调查认定的事实[：:]\s*(.{50,}?)(?=上述事实|本局认为|当事人违反|$)', text, re.DOTALL)
    if m:
        return m.group(1), True

    # Pattern 4: 020203 编码后段落（表格格式）
    m = re.search(r'020203[^\n]*\n(.{50,}?)(?=处\s*\n罚\s*\n依\s*\n据|处罚\s*依据|$)', text, re.DOTALL)
    if m:
        raw = m.group(1)
        raw = re.sub(r'\n[违法行事为实处罚依据类型]\n', '\n', raw)
        return raw, True

    # Pattern 5: 找含明确行贿关键词的长文本块
    _BRIBERY_BLOCK_RE = re.compile(r'(行贿|贿赂|好处费|回扣|带客|谋取交易|以谋取|介绍费)')
    lines = text.split('\n')
    long_blocks, current = [], []
    for line in lines:
        s = line.strip()
        if len(s) > 20:
            current.append(s)
        else:
            if current:
                block = ''.join(current)
                if len(block) > 50 and _BRIBERY_BLOCK_RE.search(block):
                    long_blocks.append(block)
                current = []
    if current:
        block = ''.join(current)
        if len(block) > 50 and _BRIBERY_BLOCK_RE.search(block):
            long_blocks.append(block)

    if long_blocks:
        best = max(long_blocks, key=lambda b: _score_sentence(b))
        return best, False

    # Pattern 6: 单行违法行为描述
    m = re.search(r'违法行为.{0,5}[：:]\s*(.{10,200})', text)
    if m:
        return m.group(1).strip()[:250], False

    return None, False


def extract_main_fact(text, company_name=''):
    """提取主要事实：优先 AI 结构化摘要，fallback 规则选句"""

    raw_block, is_structured = _find_raw_fact_block(text)

    if raw_block:
        # AI 摘要优先
        ai_result = _ai_summarize_fact(raw_block, company_name)
        if ai_result:
            return ai_result
        # 退化：规则压缩，但仍过滤法条套语
        condensed = _condense_fact(raw_block)
        if _is_law_boilerplate(condensed):
            return '（原始处罚文书未载明违法事实详情）'
        return condensed

    # 没找到结构化块时，让 AI 从全文中尝试提取
    ai_result = _ai_summarize_fact(text[:4000], company_name)
    if ai_result and ai_result != '/':
        return ai_result

    return '（原始处罚文书未载明违法事实详情）'


# ==================== 法律依据提取 ====================

def extract_legal_basis(text, punishment_date=None):
    # 宽松匹配：允许文字间有空白（OCR常见问题）
    def ws(s):
        return r'\s*'.join(s)

    explicit_2025 = bool(re.search(r'2025修订|（2025修订）', text) or
                         re.search(r'第八条[\s\S]{0,100}第二十四条', text))
    explicit_2019 = bool(re.search(r'2019修正|（2019修正）', text) or
                         re.search(r'第七条[\s\S]{0,100}第十九条', text))

    if not explicit_2025 and not explicit_2019:
        if re.search(r'第八条.{0,80}贿赂|贿赂.{0,80}第八条|第二十四条', text):
            explicit_2025 = True
        elif re.search(r'第七条.{0,80}贿赂|贿赂.{0,80}第七条|第十九条', text):
            explicit_2019 = True
        else:
            if punishment_date:
                m = re.match(r'(\d{4})[.\-](\d{1,2})', punishment_date)
                if m:
                    y, mo = int(m.group(1)), int(m.group(2))
                    explicit_2025 = (y > 2025) or (y == 2025 and mo >= 11)
                    explicit_2019 = not explicit_2025
            else:
                explicit_2025 = True

    if explicit_2025:
        ver = '（2025修订）'
        bribery_art = '第八条'
        penalty_art = '第二十四条'
    else:
        ver = '（2019修正）'
        bribery_art = '第七条'
        penalty_art = '第十九条'

    sub = ''
    m = re.search(rf'{bribery_art}第一款第[（(]?([一二三])[）)]?项', text)
    if m:
        sub = f'第一款第（{m.group(1)}）项'
    elif re.search(rf'{bribery_art}第一款', text):
        sub = '第一款'

    has_para3 = bool(re.search(rf'{bribery_art}第三款', text))

    lines = [
        f'《中华人民共和国反不正当竞争法》{ver}{bribery_art}' +
        (f'{sub}' if sub else '') +
        ('、第三款' if has_para3 else ''),
        f'《中华人民共和国反不正当竞争法》{ver}{penalty_art}',
    ]

    # 药品管理法（允许文字间空白）
    flat = re.sub(r'\s+', ' ', text)
    if re.search(r'药\s*品\s*管\s*理\s*法', flat):
        if re.search(r'第\s*八\s*十\s*八\s*条', flat):
            lines.append('《中华人民共和国药品管理法》第八十八条')
        if re.search(r'第\s*一\s*百\s*四\s*十\s*一\s*条', flat):
            lines.append('《中华人民共和国药品管理法》第一百四十一条')

    # 网络反不正当竞争暂行规定
    if re.search(r'网\s*络\s*反\s*不\s*正\s*当\s*竞\s*争', flat):
        if re.search(r'第\s*十\s*条', flat):
            lines.append('《网络反不正当竞争暂行规定》第十条')

    return '\n'.join(lines)


# ==================== PDF 处理 ====================

def process_pdf(pdf_path):
    filename = os.path.basename(pdf_path)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception as e:
        print(f'  [WARN] PDF读取失败 {filename}: {e}')
        return None

    if not text.strip():
        return None

    suspected = False
    if not is_bribery_case(text):
        if is_suspected_bribery(text):
            suspected = True
        else:
            return None

    institution = extract_institution(text)
    punishment_time = extract_punishment_time(text)
    company = extract_company_name(text, filename)

    remark = _SUSPECTED_REMARK if suspected else ''

    return {
        '处罚时间':         punishment_time,
        '案例标题':         company,
        '行贿方':           company,
        '被行贿方':         extract_bribee(text),
        '被行贿方角色':     classify_bribee_role(text),
        '行贿手段':         classify_bribe_method(text),
        '所处行业':         classify_industry(company, text),
        '主要事实':         extract_main_fact(text, company),
        '行贿金额':         extract_bribe_amount(text),
        '没收违法所得金额': extract_confiscated_amount(text),
        '罚款金额':         extract_fine_amount(text),
        '处罚机构':         institution,
        '地域':             extract_region(institution, text),
        '案号':             extract_case_number(text),
        '法律依据':         extract_legal_basis(text, punishment_time),
        '备注':             remark,
        '_source':          'wk',
        '_pdf_path':        pdf_path,
    }


# ==================== 企查查 Excel 处理 ====================

def process_qcc_excel(path):
    cases = []
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row[0]:
            continue
        try:
            _, name, case_num, date, institution, facts, result, data_type = (list(row) + [None]*8)[:8]
        except Exception:
            continue

        if not name:
            continue

        combined = f'{facts or ""} {result or ""}'

        suspected = False
        if not is_bribery_case(combined):
            if is_suspected_bribery(combined):
                suspected = True
            else:
                continue

        if str(data_type or '').strip() in ('失效', '注销'):
            continue

        if date:
            if hasattr(date, 'strftime'):
                date_str = date.strftime('%Y.%m.%d')
            else:
                date_str = normalize_date(str(date))
        else:
            date_str = '/'

        name = str(name).strip()
        institution = str(institution or '/').strip()

        remark_parts = ['来自于企查查']
        if suspected:
            remark_parts.append(_SUSPECTED_REMARK)

        cases.append({
            '处罚时间':         date_str,
            '案例标题':         name,
            '行贿方':           name,
            '被行贿方':         extract_bribee(combined),
            '被行贿方角色':     classify_bribee_role(combined),
            '行贿手段':         classify_bribe_method(combined),
            '所处行业':         classify_industry(name, combined),
            '主要事实':         _condense_fact(str(facts or '')),
            '行贿金额':         extract_bribe_amount(combined),
            '没收违法所得金额': extract_confiscated_amount(combined),
            '罚款金额':         extract_fine_amount(combined),
            '处罚机构':         institution,
            '地域':             extract_region(institution, combined),
            '案号':             str(case_num or '/').strip(),
            '法律依据':         extract_legal_basis(combined, date_str),
            '备注':             '；'.join(remark_parts),
            '_source':          'qcc',
        })

    return cases


# ==================== 去重 ====================

def _normalize_case_num(s):
    if not s or s == '/':
        return ''
    s = str(s)
    s = s.replace('〔', '[').replace('〕', ']').replace('﹝', '[').replace('﹞', ']')
    s = s.replace('（', '(').replace('）', ')')
    return re.sub(r'\s+', '', s).strip()


def deduplicate(wk_cases, qcc_cases):
    """案号为主键，威科先行优先"""
    seen = {}

    for case in wk_cases:
        key = _normalize_case_num(case.get('案号'))
        if not key:
            key = f"_{case.get('行贿方','')}__{case.get('处罚时间','')}"
        if key not in seen:
            seen[key] = case

    for case in qcc_cases:
        key = _normalize_case_num(case.get('案号'))
        if not key:
            key = f"_{case.get('行贿方','')}__{case.get('处罚时间','')}"
        if key not in seen:
            seen[key] = case

    return list(seen.values())


# ==================== Excel 输出 ====================

COLUMNS = [
    '序号', '处罚时间', '案例标题', '行贿方', '被行贿方',
    '被行贿方角色', '行贿手段', '所处行业', '主要事实',
    '行贿金额', '没收违法所得金额', '罚款金额',
    '处罚机构', '地域', '案号', '法律依据', '备注',
]

COL_WIDTHS = [6, 13, 28, 28, 28, 32, 18, 14, 65, 14, 14, 14, 28, 12, 38, 70, 28]

HEADER_FILL = PatternFill(start_color='17375E', end_color='17375E', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='微软雅黑')
BODY_FONT   = Font(size=10, name='微软雅黑')


def write_excel(cases, output_path, sheet_title='商业贿赂案例'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.freeze_panes = 'B2'

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 25

    for row_idx, case in enumerate(cases, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        for col_idx, col_name in enumerate(COLUMNS[1:], 2):
            val = case.get(col_name, '/')
            if val is None or val == '':
                val = '/'
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f'  ✓ 写出 {len(cases)} 条 → {output_path}')


# ==================== PDF 重命名 ====================

def rename_organized_pdfs(ordered_cases, organized_dir):
    """根据新 Excel 序号，对整理文件夹中的 WK PDF 加/更新序号前缀。

    匹配规则：去掉现有 "NNN_" 前缀后，与 process_pdf 记录的源文件名对比。
    QCC 案例无对应 WK 文件，自动跳过。
    """
    if not os.path.isdir(organized_dir):
        print(f'  [SKIP] 整理文件夹不存在，跳过重命名: {organized_dir}')
        return

    # 建立现有文件索引: stripped_name → full_path
    existing = {}
    for fname in os.listdir(organized_dir):
        if not fname.lower().endswith('.pdf'):
            continue
        stripped = re.sub(r'^\d{3}_', '', fname)
        existing[stripped] = os.path.join(organized_dir, fname)

    renamed = skipped = not_found = 0

    for seq, case in enumerate(ordered_cases, 1):
        if case.get('_source') != 'wk':
            continue
        pdf_path = case.get('_pdf_path', '')
        if not pdf_path:
            continue
        src_name = os.path.basename(pdf_path)
        new_name = f'{seq:03d}_{src_name}'

        if src_name not in existing:
            not_found += 1
            continue

        current_path = existing[src_name]
        current_name = os.path.basename(current_path)
        if current_name == new_name:
            skipped += 1
        else:
            new_path = os.path.join(organized_dir, new_name)
            os.rename(current_path, new_path)
            print(f'  [{seq:03d}] {current_name}  →  {new_name}')
            renamed += 1

    print(f'  重命名结果: 已更新 {renamed} 个 / 已是正确命名 {skipped} 个 / 整理文件夹中未找到 {not_found} 个')


# ==================== 主流程 ====================

def get_year(date_str):
    if not date_str or date_str == '/':
        return 0
    m = re.match(r'(\d{4})', str(date_str))
    return int(m.group(1)) if m else 0


def scan_pdf_dir_flat(directory, label):
    """扫描目录下的PDF（无子目录）"""
    pdfs = sorted(glob.glob(os.path.join(directory, '*.pdf')))
    cases = []
    for pdf_path in pdfs:
        case = process_pdf(pdf_path)
        if case:
            cases.append(case)
    print(f'  {label}: {len(pdfs)} 个PDF → 命中 {len(cases)} 条')
    return cases


def main():
    print('=' * 60)
    print('  商业贿赂案例提取器 v2.1')
    print('=' * 60)
    if USE_AI_SUMMARY:
        print('  ✦ AI摘要模式：I列由 Claude 自动生成结构化摘要（较慢）')
    else:
        print('  ✦ 规则模式：claude CLI 未找到，使用规则提取 I 列')
    print()

    # 1. 威科先行主目录（WK_* 子目录）
    print('【1/4】扫描威科先行主目录 ...')
    pdf_dirs = sorted(glob.glob(os.path.join(PDF_BASE_DIR, 'WK_*')))
    wk_cases = []
    total_pdfs = 0
    for d in pdf_dirs:
        pdfs = sorted(glob.glob(os.path.join(d, '*.pdf')))
        before = len(wk_cases)
        for pdf_path in pdfs:
            case = process_pdf(pdf_path)
            if case:
                wk_cases.append(case)
        total_pdfs += len(pdfs)
        sub = os.path.basename(d)
        print(f'  {sub}: {len(pdfs)} 个PDF → 新增 {len(wk_cases) - before} 条')
    print(f'  主目录合计: {total_pdfs} 个PDF → 命中 {len(wk_cases)} 条')
    print()

    # 2. 威科先行补充数据
    print('【2/4】扫描威科先行补充数据 ...')
    supp_cases = []
    if os.path.exists(WK_SUPP_DIR):
        supp_cases = scan_pdf_dir_flat(WK_SUPP_DIR, '补充数据')
    else:
        print(f'  [SKIP] 目录不存在: {WK_SUPP_DIR}')
    print()

    # 3. 企查查 Excel
    print('【3/4】处理企查查 Excel ...')
    qcc_files = glob.glob(os.path.join(QCC_DIR, '*.xlsx'))
    qcc_all = []
    for f in qcc_files:
        cases = process_qcc_excel(f)
        print(f'  {os.path.basename(f)}: {len(cases)} 条')
        qcc_all.extend(cases)
    print(f'  企查查合计 {len(qcc_all)} 条')
    print()

    # 4. 去重、拆分、输出
    print('【4/4】去重与输出 ...')
    all_wk = wk_cases + supp_cases
    all_cases = deduplicate(all_wk, qcc_all)
    print(f'  威科先行（含补充）: {len(all_wk)} 条')
    print(f'  企查查:             {len(qcc_all)} 条')
    print(f'  去重后共:           {len(all_cases)} 条')

    all_cases.sort(key=lambda c: (c.get('处罚时间') or ''))

    cases_2026  = [c for c in all_cases if get_year(c.get('处罚时间')) >= 2026]
    cases_prior = [c for c in all_cases if 0 < get_year(c.get('处罚时间')) < 2026]

    print(f'  2026年:      {len(cases_2026)} 条')
    print(f'  2024-2025年: {len(cases_prior)} 条')
    print()

    write_excel(cases_2026,  OUTPUT_2026,  '2026年商业贿赂案例')
    if cases_prior:
        write_excel(cases_prior, OUTPUT_PRIOR, '2024-2025年商业贿赂案例')

    # 5. 根据新 Excel 序号重命名整理文件夹中的 PDF
    print('【5/5】更新整理文件夹 PDF 序号前缀 ...')
    rename_organized_pdfs(cases_2026, ORGANIZED_DIR)

    print()
    print('✅ 完成！')
    print(f'   主文件  : {OUTPUT_2026}')
    if cases_prior:
        print(f'   往年文件: {OUTPUT_PRIOR}')
    print(f'   PDF整理 : {ORGANIZED_DIR}')


if __name__ == '__main__':
    main()
