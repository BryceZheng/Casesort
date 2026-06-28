#!/usr/bin/env python3
"""
CBcaseSum Skill - 商业贿赂案例库整理工具
功能：从 Word/Docx 案例文件中提取信息，按照规范格式整理到 Excel

使用方法：
    python3 extract_cases.py <输入目录> <输出Excel文件>

示例：
    python3 extract_cases.py "~/Desktop/商业贿赂docx" "~/Desktop/输出.xlsx"
"""

import os
import sys
import re
from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class CBcaseSumExtractor:
    """商业贿赂案例提取器"""

    # 商业贿赂关键词
    BRIBERY_KEYWORDS = ['贿赂', '行贿', '商业贿赂']

    # 省份列表（短名 → 全称）
    PROVINCES = {
        '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
        '河北': '河北省', '山西': '山西省', '辽宁': '辽宁省', '吉林': '吉林省',
        '黑龙江': '黑龙江省', '江苏': '江苏省', '浙江': '浙江省', '安徽': '安徽省',
        '福建': '福建省', '江西': '江西省', '山东': '山东省', '河南': '河南省',
        '湖北': '湖北省', '湖南': '湖南省', '广东': '广东省', '广西': '广西壮族自治区',
        '海南': '海南省', '四川': '四川省', '贵州': '贵州省', '云南': '云南省',
        '西藏': '西藏自治区', '陕西': '陕西省', '甘肃': '甘肃省',
        '青海': '青海省', '宁夏': '宁夏回族自治区', '新疆': '新疆维吾尔自治区',
        '内蒙古': '内蒙古自治区',
    }

    # 城市 → 省级单位（用于从处罚机构名称推断地域）
    CITY_TO_PROVINCE = {
        '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
        '石家庄': '河北省', '保定': '河北省', '唐山': '河北省',
        '太原': '山西省', '大同': '山西省',
        '沈阳': '辽宁省', '大连': '辽宁省',
        '长春': '吉林省', '吉林': '吉林省',
        '哈尔滨': '黑龙江省', '齐齐哈尔': '黑龙江省',
        '南京': '江苏省', '苏州': '江苏省', '无锡': '江苏省', '常州': '江苏省',
        '杭州': '浙江省', '宁波': '浙江省', '温州': '浙江省', '嘉兴': '浙江省',
        '嘉善': '浙江省', '东阳': '浙江省', '义乌': '浙江省', '金华': '浙江省',
        '台州': '浙江省', '绍兴': '浙江省', '舟山': '浙江省', '平阳': '浙江省',
        '合肥': '安徽省', '芜湖': '安徽省', '池州': '安徽省', '石台': '安徽省',
        '福州': '福建省', '厦门': '福建省', '泉州': '福建省', '漳州': '福建省',
        '宁德': '福建省', '建瓯': '福建省',
        '南昌': '江西省', '九江': '江西省', '赣州': '江西省',
        '济南': '山东省', '青岛': '山东省', '烟台': '山东省', '禹城': '山东省',
        '郑州': '河南省', '洛阳': '河南省',
        '武汉': '湖北省', '宜昌': '湖北省',
        '长沙': '湖南省', '株洲': '湖南省',
        '广州': '广东省', '深圳': '广东省', '东莞': '广东省', '中山': '广东省',
        '汕头': '广东省', '佛山': '广东省',
        '南宁': '广西壮族自治区', '桂林': '广西壮族自治区', '柳州': '广西壮族自治区',
        '临桂': '广西壮族自治区',
        '海口': '海南省', '三亚': '海南省',
        '成都': '四川省', '内江': '四川省', '绵阳': '四川省',
        '贵阳': '贵州省', '遵义': '贵州省',
        '昆明': '云南省', '大理': '云南省',
        '西安': '陕西省', '宝鸡': '陕西省',
        '兰州': '甘肃省',
        '西宁': '青海省',
        '银川': '宁夏回族自治区',
        '乌鲁木齐': '新疆维吾尔自治区',
        '呼和浩特': '内蒙古自治区',
        '蒙城': '安徽省', '宿州': '安徽省', '六安': '安徽省', '黄山': '安徽省',
        '定远': '安徽省', '望江': '安徽省', '桐城': '安徽省', '滁州': '安徽省',
        '万州': '重庆市',
        '饶平': '广东省', '龙华': '广东省', '惠州': '广东省', '珠海': '广东省',
        '常熟': '江苏省', '张家港': '江苏省', '泰州': '江苏省', '南通': '江苏省',
        '扬州': '江苏省', '镇江': '江苏省', '徐州': '江苏省', '连云港': '江苏省',
        '恩施': '湖北省', '保康': '湖北省', '襄阳': '湖北省', '荆州': '湖北省',
        '十堰': '湖北省', '黄石': '湖北省', '孝感': '湖北省',
        '祁阳': '湖南省', '沅陵': '湖南省', '永州': '湖南省', '湘潭': '湖南省',
        '常德': '湖南省', '益阳': '湖南省', '怀化': '湖南省',
        '上饶': '江西省', '泰和': '江西省', '余江': '江西省', '安义': '江西省',
        '兴国': '江西省', '靖安': '江西省', '吉安': '江西省', '萍乡': '江西省',
        '奉贤': '上海市', '闵行': '上海市', '嘉定': '上海市', '松江': '上海市',
        '广元': '四川省', '南充': '四川省', '宜宾': '四川省', '泸州': '四川省',
        '达州': '四川省', '德阳': '四川省', '乐山': '四川省',
        '河池': '广西壮族自治区', '贺州': '广西壮族自治区', '百色': '广西壮族自治区',
        '玉林': '广西壮族自治区', '钦州': '广西壮族自治区',
        '赤峰': '内蒙古自治区', '包头': '内蒙古自治区',
        '丹东': '辽宁省', '锦州': '辽宁省', '鞍山': '辽宁省',
        '鄂州': '湖北省', '荆门': '湖北省', '随州': '湖北省',
    }

    # 行业关键词（按优先级排列，第一匹配优先）
    INDUSTRY_RULES = [
        ('医疗器械', ['医疗器械', '器械公司', '器械有限', '医疗科技', '医疗设备']),
        ('医疗机构', ['医院', '诊所', '卫生所', '卫生院', '医疗中心', '门诊']),
        ('医药', ['药业', '药品', '药房', '药店', '制药', '医药', '大药房']),
        ('包装材料', ['包装材料', '包装科技', '包装有限']),
        ('印刷', ['印刷']),
        ('木业/家具', ['木业', '木材', '家具', '木制品']),
        ('金属/铝制品', ['铝', '钢铁', '金属', '五金']),
        ('建材/保温', ['保温材料', '建材', '装修', '装饰材料']),
        ('工程设计', ['设计院', '规划设计', '勘察设计', '设计有限']),
        ('工程建设', ['建设', '施工', '安装工程', '水电安装', '工程安装', '建工', '建筑']),
        ('工业设备/机械', ['数控机床', '机床', '机械设备', '数控设备', '工业设备', '机械制造']),
        ('硅胶/橡塑', ['硅胶', '橡胶', '塑料']),
        ('跨境贸易', ['跨境', '进出口']),
        ('食品/餐饮', ['食品', '饮料', '餐饮', '酒业', '酒类']),
        ('汽车/检测服务', ['年检', '检测机构', '机动车', '汽车检测', '验车']),
        ('科技/信息', ['科技', '信息技术', '软件', '网络技术', '文化传媒', '文创']),
        ('旅游/文化', ['旅游', '旅行社', '景区', '文化传播', '演出']),
        ('化工/检验', ['化工', '试剂', '检验', '化学品', '实验室']),
        ('医疗服务', ['非急救', '救护车', '转运', '康复', '护理院']),
        ('餐饮/外卖', ['外卖', '饿了么', '美团', '餐饮', '加盟', '连锁品牌', '快餐']),
        ('商贸/批发', ['商贸', '贸易', '批发']),
    ]

    def __init__(self, input_dir, output_file):
        self.input_dir = input_dir
        self.output_file = output_file
        self.cases = []

    def extract_from_file(self, filepath, case_number):
        """从单个 docx 文件中提取信息"""

        case_data = {
            '序号': case_number,
            '处罚时间': '',
            '案例标题': os.path.basename(filepath).replace('.docx', ''),
            '行贿方': '',
            '被行贿方': '',
            '被行贿方角色': '',
            '行贿手段': '',
            '所处行业': '',
            '主要事实': '',
            '行贿金额': '/',
            '没收违法所得金额': '/',
            '罚款金额': '',
            '处罚机构': '',
            '地域': '',
            '案号': '',
            '法律依据': '',
            '备注': ''
        }

        try:
            doc = Document(filepath)

            # 从表格提取信息（带错误处理）
            table_data = {}
            for table_idx, table in enumerate(doc.tables):
                try:
                    for row_idx, row in enumerate(table.rows):
                        try:
                            cells = row.cells
                            if len(cells) >= 2:
                                key = cells[0].text.strip()
                                value = cells[-1].text.strip() if len(cells) > 1 else ""
                                if key and value:
                                    table_data[key] = value
                        except ValueError as e:
                            # 处理合并单元格导致的错误
                            if "no tr above topmost tr" in str(e):
                                # 尝试用替代方法读取该行
                                try:
                                    for cell in row._element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc'):
                                        text = cell.text.strip() if cell.text else ""
                                        if text and key not in table_data:
                                            table_data[key] = text
                                            break
                                except:
                                    pass
                            continue
                except Exception:
                    # 整个表格读取失败，跳过
                    continue

            # 映射表格数据到案例数据
            self._map_table_data(case_data, table_data)

            # 判断是否为商业贿赂
            is_bribery = self._check_if_bribery(table_data)
            case_data['备注'] = '✓ 商业贿赂' if is_bribery else '✗ 无关（混淆行为/虚假宣传/其他）'

            # 生成主要事实
            if not case_data['主要事实']:
                case_data['主要事实'] = self._generate_main_fact(case_data, table_data, is_bribery)

            return case_data

        except Exception as e:
            print(f"  ✗ 处理失败: {str(e)[:50]}")
            return None

    def _map_table_data(self, case_data, table_data):
        """将表格数据映射到案例数据"""

        # 处罚时间
        if '处罚决定日期' in table_data:
            date_str = table_data['处罚决定日期'].split()[0]
            case_data['处罚时间'] = date_str.replace('-', '.')
        elif '处罚时间' in table_data:
            case_data['处罚时间'] = table_data['处罚时间']

        # 案号
        if '行政处罚决定书文号' in table_data:
            case_data['案号'] = table_data['行政处罚决定书文号'].replace('﹝', '〔').replace('﹞', '〕')
        elif '发文案号' in table_data:
            case_data['案号'] = table_data['发文案号']

        # 处罚机构
        case_data['处罚机构'] = table_data.get('处罚机构', table_data.get('处罚机关', ''))

        # 行贿方
        if '主体名称' in table_data:
            case_data['行贿方'] = table_data['主体名称']
        elif '行政相对人名称' in table_data:
            case_data['行贿方'] = table_data['行政相对人名称']

        # 地域：优先从住所，其次从处罚机构名称推断省份
        if '住所' in table_data and table_data['住所']:
            province = self._extract_province(table_data['住所'])
            if province:
                case_data['地域'] = province
        if not case_data['地域']:
            organ = table_data.get('处罚机构', table_data.get('处罚机关', ''))
            case_data['地域'] = self._extract_province_from_organ(organ)

        # 罚款金额
        if '罚款金额（万元）' in table_data:
            try:
                case_data['罚款金额'] = str(int(float(table_data['罚款金额（万元）']) * 10000))
            except:
                pass
        elif '处罚结果' in table_data:
            # 从"处罚结果"字段提取罚款金额（如"罚款10000元"）
            match = re.search(r'罚款([0-9,，]+)', table_data['处罚结果'])
            if match:
                case_data['罚款金额'] = match.group(1)
        elif '处罚内容' in table_data:
            match = re.search(r'罚款([0-9,，]+)', table_data['处罚内容'])
            if match:
                case_data['罚款金额'] = match.group(1)

        # 没收违法所得
        if '没收违法所得、没收非法财物的金额（万元）' in table_data:
            amount = table_data['没收违法所得、没收非法财物的金额（万元）']
            if amount and amount != '0.0':
                try:
                    case_data['没收违法所得金额'] = str(int(float(amount) * 10000))
                except:
                    pass

        # 法律依据：精确到条款，格式对齐模板
        basis_src = (table_data.get('处罚依据', '') or table_data.get('违法依据', ''))
        full_text = table_data.get('行政处罚决定书（全文或摘要）', '')
        if basis_src or full_text:
            case_data['法律依据'] = self._extract_legal_basis(basis_src, full_text)

        # 所处行业：先用公司名称精准判断，再用全文宽松判断
        company_name = case_data.get('行贿方', '') + case_data.get('案例标题', '')
        full_content  = (company_name +
                         table_data.get('违法事实', '') +
                         table_data.get('处罚事实', '') +
                         table_data.get('行政处罚决定书（全文或摘要）', ''))
        case_data['所处行业'] = self._detect_industry(full_content, company_name)

    def _check_if_bribery(self, table_data):
        """检查是否为商业贿赂相关案例"""
        # 检查多个关键字段，而不仅仅是处罚依据
        check_fields = [
            '处罚依据', '案件名称', '处罚名称', '案例标题',
            '行政处罚决定书（全文或摘要）', '违法行为类型',
            '违法事实', '处罚事由'
        ]

        for field in check_fields:
            field_text = table_data.get(field, '')
            if any(keyword in field_text for keyword in self.BRIBERY_KEYWORDS):
                return True

        return False

    def _extract_province(self, address):
        """从地址中提取省份"""
        for short_name, full_name in self.PROVINCES.items():
            if short_name in address:
                return full_name
        return ''

    def _extract_province_from_organ(self, organ_name):
        """从处罚机构名称推断省级行政单位"""
        # 先检查直辖市/自治区全称
        for short, full in self.PROVINCES.items():
            if short in organ_name:
                return full
        # 再用城市名反查省份
        for city, province in self.CITY_TO_PROVINCE.items():
            if city in organ_name:
                return province
        return '/'

    def _detect_industry(self, content, company_name=''):
        """从公司名称和案件内容推断所处行业。
        先只看公司名称（精准匹配），再看全文（宽松匹配）。"""
        name = company_name or content
        # 第一轮：仅凭公司名称/案例标题（短文本，误匹配少）
        for industry, keywords in self.INDUSTRY_RULES:
            if any(kw in name for kw in keywords):
                return industry
        # 第二轮：扩展到全文（行贿情节描述可能含有行业线索）
        # 此时排除"医疗机构"和"医药"类关键词，避免把向医院行贿的案例误标为医疗行业
        skip_in_fulltext = {'医疗机构', '医药'}
        for industry, keywords in self.INDUSTRY_RULES:
            if industry in skip_in_fulltext:
                continue
            if any(kw in content for kw in keywords):
                return industry
        return '其他'

    def _extract_legal_basis(self, basis_text, full_text=''):
        """提取精确法律依据，格式：违法条款\\n处罚条款（\\n药品管理法条款，如有）

        关键版本规则：
        - 2019修正版：第七条=商业贿赂，第十九条=商业贿赂处罚
        - 2025修订版：第七条=虚假宣传，第八条=商业贿赂，第二十四条=商业贿赂处罚
        - 过渡期案例：违法条款可能用2019修正第七条，处罚条款用2025修订第二十四条（均属正常）
        """
        combined = basis_text + ' ' + full_text

        # ── 违法条款 ──────────────────────────────────────────────────────────
        # 先找具体条文匹配（可能含款项）
        vio_m = re.search(r'第[七八]条(?:第[一二三四]款)?(?:第[（(][一二三四][）)]项)?', combined)
        if not vio_m:
            vio_m = re.search(r'第[78]条(?:第\d款)?(?:第[（(]\d[）)]项)?', combined)

        if vio_m:
            vio_clause = vio_m.group()
            is_art8 = '八' in vio_clause or '8' in vio_clause
        else:
            # 无法从文本提取条款时，根据版本设默认
            is_art8 = bool(re.search(r'2025.*修[订正]|修[订正].*2025|第二十四条', combined))
            vio_clause = '第八条' if is_art8 else '第七条'

        if is_art8:
            vio_law = '《中华人民共和国反不正当竞争法》（2025修订）'
        else:
            vio_law = '《中华人民共和国反不正当竞争法》（2019修正）'

        # ── 处罚条款 ──────────────────────────────────────────────────────────
        pen_m = re.search(r'第(?:十九|二十四)条(?:第[一二三四]款)?', combined)
        if pen_m:
            pen_clause = pen_m.group()
            is_pen_2025 = '二十四' in pen_clause
        else:
            is_pen_2025 = is_art8 or bool(re.search(r'2025.*修[订正]|修[订正].*2025', combined))
            pen_clause = '第二十四条' if is_pen_2025 else '第十九条'

        if is_pen_2025:
            pen_law = '《中华人民共和国反不正当竞争法》（2025修订）'
        else:
            pen_law = '《中华人民共和国反不正当竞争法》（2019修正）'

        result = f'{vio_law}{vio_clause}\n{pen_law}{pen_clause}'

        # ── 药品管理法（如有引用）────────────────────────────────────────────
        # 药品贿赂案件可能同时违反药品管理法第88条（贿赂）和第141条（处罚）
        drug_m = re.search(r'药品管理法[^，。]*第(?:八十八|141|88)条', combined)
        if drug_m:
            result += '\n《中华人民共和国药品管理法》第八十八条'
            if re.search(r'第(?:一百四十一|141)条', combined):
                result += '\n《中华人民共和国药品管理法》第一百四十一条'

        return result

    def _condense_fact(self, raw_text):
        """将原始事实文本压缩为3-4句话（约150-250字）"""
        if not raw_text or len(raw_text) < 50:
            return raw_text or '/'

        # 清理前缀
        text = re.sub(r'^(主要违法事实|违法事实)[：:]\s*', '', raw_text.strip())
        text = re.sub(r'\s+', ' ', text)

        # 按句号分句
        sentences = [s.strip() for s in re.split(r'(?<=。)', text) if s.strip()]

        # 过滤纯程序性语句（认定初次违法、无法计算违法所得等）
        SKIP_PATTERNS = ['认定当事人属', '违法所得无法计算', '无法从当事人收取', '通过.*系统查询']
        core = []
        for s in sentences:
            if any(re.search(p, s) for p in SKIP_PATTERNS):
                continue
            core.append(s)
            if len(core) >= 4:
                break

        result = ''.join(core[:4])
        # 如果结果超过300字，在第3句截断
        if len(result) > 300 and len(core) > 3:
            result = ''.join(core[:3])

        return result if result else text[:250]

    def _generate_main_fact(self, case_data, table_data, is_bribery):
        """从文档提取主要事实并压缩为3-4句话"""
        # 按优先级尝试各字段
        for field in ['违法事实', '处罚事实']:
            raw = table_data.get(field, '').strip()
            if raw:
                return self._condense_fact(raw)

        # 行政处罚决定书全文：提取主要违法事实段落
        full = table_data.get('行政处罚决定书（全文或摘要）', '').strip()
        if full:
            m = re.search(
                r'主要违法事实[：:]\s*(.+?)(?:行政处罚种类|处罚依据|行政处罚决定|$)',
                full, re.DOTALL
            )
            raw = m.group(1) if m else full
            return self._condense_fact(raw)

        # 文档中无事实描述（摘要格式），生成标准说明
        if is_bribery:
            company = case_data.get('行贿方', '当事人')
            return f"（原始处罚文书未载明违法事实详情）{company}涉嫌商业贿赂，违反《反不正当竞争法》相关规定。"
        else:
            violation = table_data.get('违法行为类型', '')
            company = case_data.get('行贿方', '当事人')
            if '混淆' in violation or '地理标志' in violation:
                return f"{company}销售产品时使用受保护的地理标志或进行混淆行为。"
            elif '虚假' in violation or '广告' in violation:
                return f"{company}涉及虚假宣传或违法广告。"
            else:
                return f"{company}违反反不正当竞争法的相关规定。"

    def process_directory(self):
        """处理整个目录"""
        docx_files = sorted([f for f in os.listdir(self.input_dir) if f.endswith('.docx')])

        print(f"开始处理 {len(docx_files)} 个案例文件...\n")

        case_number = 1
        for idx, filename in enumerate(docx_files, 1):
            filepath = os.path.join(self.input_dir, filename)
            print(f"{idx:3d}. {filename[:50]:50} ", end="", flush=True)

            case_data = self.extract_from_file(filepath, case_number)

            if case_data:
                self.cases.append(case_data)
                status = '✓' if '商业贿赂' in case_data['备注'] else '✗'
                print(f"{status}")
                case_number += 1
            else:
                print("✗")

        print(f"\n处理完成！共提取 {len(self.cases)} 个案例\n")

    def save_to_excel(self):
        """保存到 Excel 文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2026"

        headers = [
            '序号', '处罚时间', '案例标题', '行贿方', '被行贿方', '被行贿方角色',
            '行贿手段', '所处行业', '主要事实', '行贿金额', '没收违法所得金额',
            '罚款金额', '处罚机构', '地域', '案号', '法律依据', '备注'
        ]

        # 添加标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 设置列宽
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['I'].width = 35
        ws.column_dimensions['O'].width = 20

        # 添加数据行
        for row_idx, case in enumerate(self.cases, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = case.get(header, '')
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # 背景色
                if '商业贿赂' in str(case.get('备注', '')):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb.save(self.output_file)
        print(f"✅ Excel 已保存: {self.output_file}")

    def print_summary(self):
        """打印统计摘要"""
        bribery_count = sum(1 for c in self.cases if '商业贿赂' in c['备注'])
        print("="*80)
        print("案例统计摘要:")
        print("="*80)
        print(f"✓ 商业贿赂相关: {bribery_count} 个")
        print(f"✗ 无关案例: {len(self.cases) - bribery_count} 个")
        print(f"总计: {len(self.cases)} 个\n")


def main():
    if len(sys.argv) < 3:
        print("使用方法: python3 extract_cases.py <输入目录> <输出Excel文件>")
        print("示例: python3 extract_cases.py '/path/to/docx' '/path/to/output.xlsx'")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_file = sys.argv[2]

    if not os.path.isdir(input_dir):
        print(f"错误：输入目录不存在: {input_dir}")
        sys.exit(1)

    extractor = CBcaseSumExtractor(input_dir, output_file)
    extractor.process_directory()
    extractor.save_to_excel()
    extractor.print_summary()


if __name__ == '__main__':
    main()
