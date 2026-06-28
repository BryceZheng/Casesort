#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
第二阶段（人工核验版）：
  从人工核验后的文件夹直接提取结构化字段，生成案例库 Excel。

  假定条件：
    - 文件夹内的 PDF 已经人工确认为商业贿赂案例，不做贿赂过滤
    - 企查查数据放在 <INPUT_DIR>/企查查数据/ 子目录下，格式与原始 QCC 导出一致
    - WK PDF 与 QCC 重叠时，只保留一条（WK 优先），QCC 专属案例在备注注明

  输出：
    - 2026年版    → OUTPUT_2026
    - 2024-2025年 → OUTPUT_PRIOR（有案例时才生成）
"""

import os, re, glob, sys
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 从同目录的主脚本导入所有提取函数
_SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _SKILL_DIR)
import extract_bribery_cases as _base

# ==================== 路径配置 ====================
INPUT_DIR    = os.path.expanduser('~/Desktop/人工核验后的商业贿赂案例')
QCC_SUBDIR   = '企查查数据'
OUTPUT_2026  = os.path.expanduser('~/Desktop/商业贿赂案例库2026年版.xlsx')
OUTPUT_PRIOR = os.path.expanduser('~/Desktop/商业贿赂案例库(2024-2025年).xlsx')

_QCC_ONLY_NOTE = '仅见于企查查（威科先行未收录）'


# ==================== PDF 提取（不过滤，直接提取所有字段） ====================

def process_pdf_all(pdf_path):
    """提取字段，不做贿赂过滤（用户已人工确认）。"""
    filename = os.path.basename(pdf_path)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception as e:
        print(f'  [WARN] 读取失败 {filename}: {e}')
        return None

    if not text.strip():
        print(f'  [WARN] 文本为空，已跳过: {filename}')
        return None

    institution     = _base.extract_institution(text)
    punishment_time = _base.extract_punishment_time(text)
    company         = _base.extract_company_name(text, filename)

    return {
        '处罚时间':         punishment_time,
        '案例标题':         company,
        '行贿方':           company,
        '被行贿方':         _base.extract_bribee(text),
        '被行贿方角色':     _base.classify_bribee_role(text),
        '行贿手段':         _base.classify_bribe_method(text),
        '所处行业':         _base.classify_industry(company, text),
        '主要事实':         _base.extract_main_fact(text, company),
        '行贿金额':         _base.extract_bribe_amount(text),
        '没收违法所得金额': _base.extract_confiscated_amount(text),
        '罚款金额':         _base.extract_fine_amount(text),
        '处罚机构':         institution,
        '地域':             _base.extract_region(institution, text),
        '案号':             _base.extract_case_number(text),
        '法律依据':         _base.extract_legal_basis(text, punishment_time),
        '备注':             '',
        '_source':          'wk',
        '_pdf_path':        pdf_path,
    }


# ==================== QCC Excel 提取（不过滤，全量读取） ====================

def _norm_case_num(s):
    if not s or str(s).strip() == '/':
        return ''
    s = str(s)
    for a, b in [('〔','['),('〕',']'),('﹝','['),('﹞',']'),('（','('),('）',')')]:
        s = s.replace(a, b)
    return re.sub(r'\s+', '', s).strip()


def process_qcc_excel_all(path):
    """读取 QCC Excel，不做贿赂过滤（用户已人工确认该文件）。"""
    cases = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f'  [WARN] 读取 Excel 失败 {os.path.basename(path)}: {e}')
        return cases

    ws = wb.active

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        try:
            _, name, case_num, date, institution, facts, result, data_type = (list(row) + [None]*8)[:8]
        except Exception:
            continue

        if not name:
            continue

        # 只排除明确无效状态
        if str(data_type or '').strip() in ('失效', '注销'):
            continue

        if date:
            if hasattr(date, 'strftime'):
                date_str = date.strftime('%Y.%m.%d')
            else:
                date_str = _base.normalize_date(str(date))
        else:
            date_str = '/'

        name        = str(name).strip()
        institution = str(institution or '/').strip()
        combined    = f'{facts or ""} {result or ""}'

        cases.append({
            '处罚时间':         date_str,
            '案例标题':         name,
            '行贿方':           name,
            '被行贿方':         _base.extract_bribee(combined),
            '被行贿方角色':     _base.classify_bribee_role(combined),
            '行贿手段':         _base.classify_bribe_method(combined),
            '所处行业':         _base.classify_industry(name, combined),
            '主要事实':         _base.extract_main_fact(combined, name),
            '行贿金额':         _base.extract_bribe_amount(combined),
            '没收违法所得金额': _base.extract_confiscated_amount(combined),
            '罚款金额':         _base.extract_fine_amount(combined),
            '处罚机构':         institution,
            '地域':             _base.extract_region(institution, combined),
            '案号':             str(case_num or '/').strip(),
            '法律依据':         _base.extract_legal_basis(combined, date_str),
            '备注':             '',
            '_source':          'qcc',
        })

    return cases


# ==================== 去重 + QCC 专属标注 ====================

def deduplicate_and_mark(wk_cases, qcc_cases):
    """WK 优先去重；仅出现在 QCC 的案例在备注中注明。"""
    seen = {}

    for case in wk_cases:
        key = _norm_case_num(case.get('案号')) or \
              f"__{case.get('行贿方','')}__{case.get('处罚时间','')}"
        if key not in seen:
            seen[key] = case

    qcc_only_count = 0
    for case in qcc_cases:
        key = _norm_case_num(case.get('案号')) or \
              f"__{case.get('行贿方','')}__{case.get('处罚时间','')}"
        if key not in seen:
            case = dict(case)
            existing = (case.get('备注') or '').strip()
            if _QCC_ONLY_NOTE not in existing:
                case['备注'] = (_QCC_ONLY_NOTE + '；' + existing).rstrip('；') \
                               if existing else _QCC_ONLY_NOTE
            seen[key] = case
            qcc_only_count += 1

    return list(seen.values()), qcc_only_count


# ==================== 主流程 ====================

def main():
    print('=' * 60)
    print('  商业贿赂案例提取器（人工核验版）')
    print('=' * 60)
    if _base.USE_AI_SUMMARY:
        print('  ✦ AI摘要模式：主要事实由 Claude 生成（较慢，约 5s/条）')
    else:
        print('  ✦ 规则模式：claude CLI 未找到，使用规则提取主要事实')
    print(f'  输入目录: {INPUT_DIR}')
    print()

    # 1. 扫描 WK PDF
    pdf_files = sorted(glob.glob(os.path.join(INPUT_DIR, '*.pdf')))
    print(f'【1/3】扫描 WK PDF（共 {len(pdf_files)} 个）...')
    wk_cases = []
    for i, pdf_path in enumerate(pdf_files, 1):
        fname = os.path.basename(pdf_path)
        print(f'  [{i:03d}/{len(pdf_files)}] {fname}')
        case = process_pdf_all(pdf_path)
        if case:
            wk_cases.append(case)
    print(f'  ✓ WK 提取完成: {len(wk_cases)} 条\n')

    # 2. 处理企查查 Excel
    qcc_dir   = os.path.join(INPUT_DIR, QCC_SUBDIR)
    qcc_files = sorted(glob.glob(os.path.join(qcc_dir, '*.xlsx'))) if os.path.isdir(qcc_dir) else []
    print(f'【2/3】处理企查查 Excel（共 {len(qcc_files)} 个）...')
    qcc_cases = []
    for f in qcc_files:
        cases = process_qcc_excel_all(f)
        print(f'  {os.path.basename(f)}: {len(cases)} 条')
        qcc_cases.extend(cases)
    print(f'  ✓ QCC 合计: {len(qcc_cases)} 条\n')

    # 3. 去重、年份拆分、输出
    print('【3/3】去重与输出 ...')
    all_cases, qcc_only_count = deduplicate_and_mark(wk_cases, qcc_cases)
    print(f'  WK 来源:         {len(wk_cases)} 条')
    print(f'  QCC 来源:        {len(qcc_cases)} 条（其中仅QCC专属 {qcc_only_count} 条）')
    print(f'  去重后合计:      {len(all_cases)} 条')

    all_cases.sort(key=lambda c: (c.get('处罚时间') or ''))

    cases_2026  = [c for c in all_cases if _base.get_year(c.get('处罚时间')) >= 2026]
    cases_prior = [c for c in all_cases if 0 < _base.get_year(c.get('处罚时间')) < 2026]
    cases_unknown = [c for c in all_cases if _base.get_year(c.get('处罚时间')) == 0]

    print(f'  2026年：         {len(cases_2026)} 条')
    print(f'  2024-2025年：    {len(cases_prior)} 条')
    if cases_unknown:
        print(f'  日期未知：       {len(cases_unknown)} 条（归入2026年文件）')
    print()

    # 日期未知的并入2026年文件
    _base.write_excel(cases_2026 + cases_unknown, OUTPUT_2026, '2026年商业贿赂案例')

    if cases_prior:
        _base.write_excel(cases_prior, OUTPUT_PRIOR, '2024-2025年商业贿赂案例')
    else:
        print('  (无2024-2025年案例，不生成往年文件)')

    print()
    print('✅ 完成！')
    print(f'   2026年版  : {OUTPUT_2026}')
    if cases_prior:
        print(f'   往年文件  : {OUTPUT_PRIOR}')


if __name__ == '__main__':
    main()
