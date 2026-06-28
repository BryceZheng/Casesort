#!/usr/bin/env python3
"""
Casesort Phase 3 - 全量核查 Excel 生成工具

用法：
  追加批次数据：
    python3 phase3_excel.py --append '<JSON>' --tmp /tmp/casesort_phase3_tmp.json

  合并输出最终 Excel：
    python3 phase3_excel.py --merge output.xlsx --tmp /tmp/casesort_phase3_tmp.json
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl：pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

COLOR_BLUE   = "BDD7EE"   # 确认纳入
COLOR_ORANGE = "F4B942"   # 疑似
COLOR_RED    = "FF7B7B"   # 确认排除
COLOR_HEADER = "2F5496"   # 表头

VERDICT_ORDER = {"确认纳入": 0, "疑似": 1, "确认排除": 2}

COLUMNS = ["案号", "案件事实", "判断结论", "判断依据"]
COL_WIDTHS = [28, 60, 12, 50]


def load_tmp(tmp_path: str) -> list:
    if os.path.exists(tmp_path):
        with open(tmp_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_tmp(tmp_path: str, data: list):
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def append_batch(batch_json: str, tmp_path: str):
    batch = json.loads(batch_json)
    existing = load_tmp(tmp_path)
    existing.extend(batch)
    save_tmp(tmp_path, existing)
    counts = {}
    for item in batch:
        v = item.get("判断结论", "未知")
        counts[v] = counts.get(v, 0) + 1
    parts = [f"{v}：{n} 个" for v, n in counts.items()]
    print(f"✅ 已追加 {len(batch)} 条记录。{' | '.join(parts)}")
    print(f"   临时文件：{tmp_path}（累计 {len(existing)} 条）")


def build_excel(data: list, output_path: str):
    # 按结论排序：确认纳入 → 疑似 → 确认排除
    data_sorted = sorted(data, key=lambda x: VERDICT_ORDER.get(x.get("判断结论", ""), 99))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全量核查报告"

    # 表头样式
    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    # 颜色映射
    fill_map = {
        "确认纳入": PatternFill(fill_type="solid", fgColor=COLOR_BLUE),
        "疑似":     PatternFill(fill_type="solid", fgColor=COLOR_ORANGE),
        "确认排除": PatternFill(fill_type="solid", fgColor=COLOR_RED),
    }

    # 写数据
    for row_idx, item in enumerate(data_sorted, 2):
        verdict = item.get("判断结论", "")
        row_fill = fill_map.get(verdict, PatternFill())

        values = [
            item.get("案号", ""),
            item.get("案件事实", ""),
            verdict,
            item.get("判断依据", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            cell.font = Font(size=10)

        # 案件事实行高（多行内容）
        lines = str(item.get("案件事实", "")).count("\n") + 1
        ws.row_dimensions[row_idx].height = max(30, lines * 15)

    # 列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 统计 sheet
    ws_stat = wb.create_sheet("统计")
    verdicts = [item.get("判断结论", "未知") for item in data]
    stat_data = {
        "确认纳入": verdicts.count("确认纳入"),
        "疑似":     verdicts.count("疑似"),
        "确认排除": verdicts.count("确认排除"),
        "合计":     len(data),
    }
    ws_stat.append(["类别", "数量"])
    for k, v in stat_data.items():
        ws_stat.append([k, v])
    ws_stat.column_dimensions["A"].width = 15
    ws_stat.column_dimensions["B"].width = 10

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    print(f"\n✅ 全量核查 Excel 已生成：{output_path}")
    print(f"   确认纳入：{stat_data['确认纳入']} | 疑似：{stat_data['疑似']} | 确认排除：{stat_data['确认排除']} | 合计：{stat_data['合计']}")


def merge(output_path: str, tmp_path: str):
    data = load_tmp(tmp_path)
    if not data:
        print("错误：临时文件为空或不存在，请先执行 --append", file=sys.stderr)
        sys.exit(1)
    build_excel(data, output_path)


def main():
    parser = argparse.ArgumentParser(description="Casesort Phase 3 全量核查 Excel 生成")
    parser.add_argument("--append", metavar="JSON", help="追加一批核查结果（JSON 字符串）")
    parser.add_argument("--merge", metavar="OUTPUT", help="合并所有批次，输出最终 Excel")
    parser.add_argument("--tmp", default="/tmp/casesort_phase3_tmp.json", help="临时存储文件路径")
    args = parser.parse_args()

    if args.append:
        append_batch(args.append, args.tmp)
    elif args.merge:
        merge(args.merge, args.tmp)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
