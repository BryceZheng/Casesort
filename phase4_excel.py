#!/usr/bin/env python3
"""
Casesort Phase 4 - 成果产出 Excel 生成工具

用法：
  追加批次数据（格式A）：
    python3 phase4_excel.py --format A --append '<JSON>' --tmp /tmp/casesort_phase4_tmp.json

  追加批次数据（格式B，自定义字段）：
    python3 phase4_excel.py --format B --fields '["字段1","字段2"]' --append '<JSON>' --tmp /tmp/casesort_phase4_tmp.json

  合并输出最终 Excel：
    python3 phase4_excel.py --merge output.xlsx --tmp /tmp/casesort_phase4_tmp.json
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl：pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

COLOR_HEADER = "2F5496"
COLOR_ROW_ODD  = "EBF3FB"
COLOR_ROW_EVEN = "FFFFFF"

# 格式A：商业贿赂行政处罚案例 17 字段
FORMAT_A_COLUMNS = [
    "序号", "处罚时间", "案例标题", "行贿方", "被行贿方",
    "被行贿方角色", "行贿手段", "所处行业", "主要事实",
    "行贿金额", "没收违法所得金额", "罚款金额",
    "处罚机构", "地域", "案号", "法律依据", "备注"
]

FORMAT_A_WIDTHS = [6, 14, 30, 20, 20, 22, 18, 14, 55, 12, 16, 12, 28, 12, 30, 50, 20]

# 格式B：通用类型案例基础字段
FORMAT_B_BASE_COLUMNS = [
    "序号", "处罚/裁判时间", "案例标题", "行为主体", "行为对方",
    "主要事实", "地域", "案号", "法律依据", "备注"
]

FORMAT_B_BASE_WIDTHS = [6, 14, 30, 22, 22, 55, 12, 30, 50, 20]


def load_tmp(tmp_path: str) -> dict:
    if os.path.exists(tmp_path):
        with open(tmp_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"format": None, "columns": [], "widths": [], "rows": []}


def save_tmp(tmp_path: str, data: dict):
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_columns_and_widths(fmt: str, extra_fields: list = None) -> tuple:
    if fmt == "A":
        return FORMAT_A_COLUMNS[:], FORMAT_A_WIDTHS[:]
    else:
        cols = FORMAT_B_BASE_COLUMNS[:]
        widths = FORMAT_B_BASE_WIDTHS[:]
        if extra_fields:
            for field in extra_fields:
                if field not in cols:
                    cols.insert(-1, field)  # 插入到"备注"前
                    widths.insert(-1, 18)
        return cols, widths


def append_batch(fmt: str, batch_json: str, tmp_path: str, extra_fields: list = None):
    batch = json.loads(batch_json)
    state = load_tmp(tmp_path)

    # 初始化或校验格式
    if state["format"] is None:
        cols, widths = get_columns_and_widths(fmt, extra_fields)
        state["format"] = fmt
        state["columns"] = cols
        state["widths"] = widths
        state["rows"] = []
    elif state["format"] != fmt:
        print(f"错误：格式不一致（已有 {state['format']}，本次传入 {fmt}）", file=sys.stderr)
        sys.exit(1)

    state["rows"].extend(batch)
    save_tmp(tmp_path, state)
    print(f"✅ 已追加 {len(batch)} 条记录（累计 {len(state['rows'])} 条）")
    print(f"   临时文件：{tmp_path}")


def build_excel(state: dict, output_path: str):
    columns = state["columns"]
    widths = state["widths"]
    rows = state["rows"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "案例库"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头
    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # 数据行
    fill_odd  = PatternFill(fill_type="solid", fgColor=COLOR_ROW_ODD)
    fill_even = PatternFill(fill_type="solid", fgColor=COLOR_ROW_EVEN)

    for row_idx, item in enumerate(rows, 2):
        # 自动填充序号
        item["序号"] = row_idx - 1

        row_fill = fill_odd if (row_idx % 2 == 0) else fill_even

        for col_idx, col_name in enumerate(columns, 1):
            val = item.get(col_name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if not isinstance(val, (int, float)) else val)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(size=10)

            # 主要事实、法律依据：顶部对齐+换行
            if col_name in ("主要事实", "法律依据", "案件事实", "判断依据"):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col_name == "序号":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        # 行高：主要事实可能多行
        fact = str(item.get("主要事实", "") or item.get("案件事实", ""))
        lines = fact.count("\n") + 1
        ws.row_dimensions[row_idx].height = max(25, min(lines * 15, 120))

    # 列宽
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行首列（序号列）
    ws.freeze_panes = "C2"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ 案例库 Excel 已生成：{output_path}")
    print(f"   共 {len(rows)} 条案例，{len(columns)} 个字段")


def merge(output_path: str, tmp_path: str):
    state = load_tmp(tmp_path)
    if not state.get("rows"):
        print("错误：临时文件为空或不存在，请先执行 --append", file=sys.stderr)
        sys.exit(1)
    build_excel(state, output_path)


def main():
    parser = argparse.ArgumentParser(description="Casesort Phase 4 成果产出 Excel 生成")
    parser.add_argument("--format", choices=["A", "B"], help="输出格式：A=商业贿赂17字段，B=通用类型")
    parser.add_argument("--fields", metavar="JSON", help="格式B的额外字段列表（JSON数组）")
    parser.add_argument("--append", metavar="JSON", help="追加一批案例数据（JSON 字符串）")
    parser.add_argument("--merge", metavar="OUTPUT", help="合并所有批次，输出最终 Excel")
    parser.add_argument("--tmp", default="/tmp/casesort_phase4_tmp.json", help="临时存储文件路径")
    args = parser.parse_args()

    if args.append:
        if not args.format:
            print("错误：追加数据时必须指定 --format", file=sys.stderr)
            sys.exit(1)
        extra = json.loads(args.fields) if args.fields else None
        append_batch(args.format, args.append, args.tmp, extra)
    elif args.merge:
        merge(args.merge, args.tmp)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
